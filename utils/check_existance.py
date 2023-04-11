import requests
import openpyxl
import datetime

path = "exel_files/ilhom.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

row = sheet_obj.max_row
column = sheet_obj.max_column

print("Total Rows:", row)
print("Total Columns:", column)

for i in range(295, row + 1):
    fullname = str(sheet_obj.cell(row=i, column=1).value).upper()

    res = requests.get('http://localhost:8000/apps/student/list/?search={}'.format(fullname))

    if len(res.json()) == 1:
        print('passed')
    else:
        student = {
            'fullname': str(sheet_obj.cell(row=i, column=1).value).upper() if sheet_obj.cell(row=i,
                                                                                             column=1).value else None,
            'gender': str(sheet_obj.cell(row=i, column=7).value).upper() if sheet_obj.cell(row=i,
                                                                                           column=7).value else None,
            'date_of_birth': sheet_obj.cell(row=i, column=11).value.date() if sheet_obj.cell(row=i,
                                                                                             column=11).value else None,
            'region_of_birth': str(sheet_obj.cell(row=i, column=12).value).upper() if sheet_obj.cell(row=i,
                                                                                                     column=12).value else None,
            'nation': str(sheet_obj.cell(row=i, column=14).value).upper() if sheet_obj.cell(row=i,
                                                                                            column=14).value else None,
            'type_of_school': str(sheet_obj.cell(row=i, column=13).value).upper() if sheet_obj.cell(row=i,
                                                                                                    column=13).value else None,
            'description': str(sheet_obj.cell(row=i, column=15).value) if sheet_obj.cell(row=i,
                                                                                         column=15).value else '',
            'from_course_to_course': str(sheet_obj.cell(row=i, column=17).value) if not sheet_obj.cell(row=i,
                                                                                                       column=17).value or sheet_obj.cell(
                row=i, column=17).value != "-" else '',

            'ID_CARD': str(sheet_obj.cell(row=i, column=8).value) if sheet_obj.cell(row=i, column=8).value else None,
            'passport_series': str(sheet_obj.cell(row=i, column=10).value) if sheet_obj.cell(row=i,
                                                                                             column=10).value else None,
            'JSHIR': None,

            'group': requests.get('http://localhost:8000/apps/group/list/?name={}'.format(
                str(sheet_obj.cell(row=i, column=5).value))).json()[0]['id'] if sheet_obj.cell(row=i,
                                                                                               column=8).value else None,
            'order_date': datetime.date(
                int(f"20{str(sheet_obj.cell(row=i, column=2).value)[6:9]}"),
                int(str(sheet_obj.cell(row=i, column=2).value)[3:5]),
                int(str(sheet_obj.cell(row=i, column=2).value)[0:2])
            ) if sheet_obj.cell(row=i, column=2).value else None,
        }

        res = requests.post("http://127.0.0.1:8000/apps/student/create-exel/", data=student)

        if res.status_code != 201:
            print("Error in this line: {}, {}".format(i, sheet_obj.cell(row=i, column=1).value).upper())
        else:
            print('Passed +')
